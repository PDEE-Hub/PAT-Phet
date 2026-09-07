#!/usr/bin/env python3
"""
extract_data.py — regenerate mou_data.js from MOU69_Claude.xlsx
Source of truth: ../MOU69_Claude.xlsx (10-sheet restructured workbook, 2026-08-27)

Run: python3 extract_data.py
Output: mou_data.js (embedded into index.html manually, or loaded via <script src>)

V1 RULE (per implementation brief): seed ONLY Q1+Q2 real data.
Q3/Q4 exist in the source sheets (workbook already computed them) but MUST NOT
be preloaded — V1 leaves them null/"not_evaluated" so the admin can practice
entering data through the Data Entry UI. This script deliberately drops any
quarter-3/quarter-4 columns it reads.
"""
import json
import openpyxl
from datetime import datetime

SRC = "../MOU69_Claude.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)

def cell(ws, r, c):
    v = ws.cell(r, c).value
    if isinstance(v, str):
        v = v.strip()
    return v

# ── 1. Master Data: KPI catalog (structure, weights, owners, thresholds) ──
# Known source quirk: this sheet lists the 2.7.3 sub-breakdown as '2.7.7.3.1'/
# '2.7.7.3.2' at the wrong nesting level (sibling of 2.7.1-2.7.4, not child of
# 2.7.3). Sheets 2/4/6 use the correct '2.7.3.1'/'2.7.3.2' nested under 2.7.3.
# We alias here so the app has one consistent ID everywhere.
ID_ALIAS = {'2.7.7.3.1': '2.7.3.1', '2.7.7.3.2': '2.7.3.2'}

ws = wb['1.Master Data']
kpis = {}
group_map = {
    'ยุทธศาสตร์': 'h1',
    'การเงิน': 'h2fin',
    'ประสิทธิภาพการดำเนินงาน': 'h2ops',
    'โครงการสำคัญ': 'h2proj',
    'บริหารสัญญา': 'h2contract',
}
current_main = None
current_group = None
for r in range(5, 47):
    main = cell(ws, r, 4)
    sub = cell(ws, r, 5)
    kpi_type_label = cell(ws, r, 6)
    name = cell(ws, r, 7)
    unit = cell(ws, r, 8)
    weight = cell(ws, r, 9)
    if kpi_type_label in ('Core Business Enablers',) or (main and str(main).startswith('3.')):
        continue  # Enablers dims (3.x) handled by separate CBE system — not in MOU V1 scope
    if main is None and sub is None:
        continue  # section header / subtotal row (no weight/name at leaf grain)
    if main is not None:
        # parent-level row (e.g. 1.1, 2.1, 2.5, 2.7, 2.8) — also a displayable rollup KPI
        current_main = str(main)
        key = current_main
        parent = None
    else:
        raw_sub = str(sub)
        key = ID_ALIAS.get(raw_sub, raw_sub)
        # 2.7.3.1 / 2.7.3.2 nest under 2.7.3, not directly under 2.7
        parent = '2.7.3' if key in ('2.7.3.1', '2.7.3.2') else current_main
    if key in kpis:
        continue
    thresholds_raw = [cell(ws, r, c) for c in range(21, 26)]
    higher_is_better = key != '1.4'
    owners_watch = [cell(ws, r, c) for c in [10] if cell(ws, r, c)]
    owners_main = [cell(ws, r, c) for c in [11, 12, 13, 14, 15, 16] if cell(ws, r, c)]
    owners_support = [cell(ws, r, c) for c in [17, 18, 19, 20] if cell(ws, r, c)]
    target_val = cell(ws, r, 37)
    target_score = cell(ws, r, 38)  # AL: "Target > คะแนน" — Excel's own pre-computed target score
    kpis[key] = {
        'id': key,
        'parent': parent,
        'label': name,
        'unit': unit,
        'weight': weight,
        'group': group_map.get(kpi_type_label, 'h2'),
        'groupLabel': kpi_type_label,
        'ownerWatch': owners_watch,
        'ownerMain': owners_main,
        'ownerSupport': owners_support,
        'thresholds': thresholds_raw,
        'higherIsBetter': higher_is_better,
        'target': target_val,
        'targetScore': target_score,
    }

# isLeaf = nothing else declares this key as its parent (true leaf of the tree,
# regardless of whether it also happens to sit at the top level, e.g. 1.2/2.6)
_parent_keys = {v['parent'] for v in kpis.values() if v['parent']}
for k, v in kpis.items():
    v['isLeaf'] = k not in _parent_keys

# ── KPI type + scoring method (per V1 brief Section 5 taxonomy) ──
# scoringMethod drives the score engine:
#   linear          — 5 numeric thresholds, linear-interpolate (official MOU rule)
#   qualitative     — admin selects Level 1-5 directly against text criteria
#   milestone_pct   — 5-step % thresholds (60/70/80/90/100), linear-interpolate
#   milestone_manual— irregular multi-stage milestone list, admin sets Level
#                     (flagged needsConfirmation — do not auto-derive, no rule in source)
#   evidence        — MILESTONE_EVIDENCE: admin sets Level from Action Plan/evidence review
KPI_TYPE = {
    '1.1': 'COMPOSITE_NUMERIC', '1.1.1': 'NUMERIC', '1.1.2': 'NUMERIC',
    '1.2': 'NUMERIC', '1.4': 'NUMERIC',
    '1.3': 'HYBRID',
    '2.1': 'COMPOSITE_NUMERIC', '2.1.1': 'NUMERIC', '2.1.2': 'NUMERIC', '2.1.3': 'NUMERIC',
    '2.2': 'NUMERIC', '2.3': 'NUMERIC', '2.4': 'NUMERIC', '2.6': 'NUMERIC',
    '2.5': 'COMPOSITE_NUMERIC', '2.5.1': 'HYBRID', '2.5.2': 'HYBRID',
    '2.7': 'PROJECT', '2.7.1': 'PROJECT', '2.7.2': 'PROJECT',
    '2.7.3': 'PROJECT', '2.7.3.1': 'PROJECT', '2.7.3.2': 'PROJECT', '2.7.4': 'PROJECT',
    '2.8': 'MILESTONE_EVIDENCE', '2.8.1': 'MILESTONE_EVIDENCE',
    '2.8.2': 'MILESTONE_EVIDENCE', '2.8.3': 'MILESTONE_EVIDENCE',
}
SCORING_METHOD = {
    '1.1.1': 'linear', '1.1.2': 'linear', '1.2': 'linear',
    '2.1.1': 'linear', '2.1.2': 'linear', '2.1.3': 'linear',
    '2.2': 'linear', '2.3': 'linear', '2.4': 'linear', '2.6': 'linear',
    '1.4': 'annual_only',
    '1.3': 'qualitative', '2.5.1': 'qualitative', '2.5.2': 'qualitative',
    '2.7.1': 'milestone_pct', '2.7.2': 'milestone_pct', '2.7.4': 'milestone_pct',
    '2.7.3.1': 'milestone_manual', '2.7.3.2': 'milestone_manual',
    '2.8.1': 'evidence', '2.8.2': 'evidence', '2.8.3': 'evidence',
}
MILESTONE_PCT_THRESHOLDS = [0.60, 0.70, 0.80, 0.90, 1.00]  # fraction, matches actual's own 0-1 scale

for k, v in kpis.items():
    v['kpiType'] = KPI_TYPE.get(k, 'NUMERIC' if v['isLeaf'] else 'COMPOSITE_NUMERIC')
    if k in SCORING_METHOD:
        v['scoringMethod'] = SCORING_METHOD[k]
    if k in ('2.7.1', '2.7.2', '2.7.4'):
        v['thresholds'] = MILESTONE_PCT_THRESHOLDS
    if k in ('2.7.3.1', '2.7.3.2'):
        v['needsConfirmation'] = True
        v['confirmationNote'] = ('มาตราวัด 8 ขั้นตอนจัดซื้อจัดจ้าง (ร้อยละ 15/20/45/50/55/60/70/100) '
                                  'ไม่ใช่เกณฑ์ 5 ระดับปกติ — ห้ามให้ระบบ auto-interpolate เป็น Level, '
                                  'ต้องให้ Admin ยืนยัน Level เองทุกไตรมาส')
    if k == '1.4':
        v['needsConfirmation'] = True
        v['confirmationNote'] = ('พบจากข้อมูลจริง: Q1/Q2/Q3 คะแนนคงที่ =1 ทั้งที่ผลสะสมต่างกันมาก '
                                  '(1.6M/3.06M/4.1M Ton เทียบเกณฑ์ ~6.01M Ton) — ตัวชี้วัดนี้ดูเหมือนจะตัดสิน '
                                  'คะแนนจริงเฉพาะ Q4 (ผลสิ้นปี) เท่านั้น ไตรมาสอื่นเป็นค่า placeholder '
                                  'V1 จึงใช้ Level ที่ Admin ยืนยัน/ค่าจาก Excel ตรงๆ ไม่ interpolate '
                                  'จนกว่าจะถึง Q4 — โปรดยืนยันกับผู้รับผิดชอบตัวชี้วัดนี้')

# ── 2. ผลรายตัวชี้วัด: leaf-level Q1/Q2 actual+score (ground truth, drop Q3/Q4) ──
ws2 = wb['2.ผลรายตัวชี้วัด']
quarterly = {}
r = 4
QMAP = {'ไตรมาส 1': 'q1', 'ไตรมาส 2': 'q2', 'ไตรมาส 3': 'q3', 'ไตรมาส 4': 'q4'}
for r in range(4, ws2.max_row + 1):
    main = cell(ws2, r, 2)
    sub = cell(ws2, r, 3)
    subsub = cell(ws2, r, 4)
    q_label = cell(ws2, r, 5)
    actual = cell(ws2, r, 7)
    score = cell(ws2, r, 8)
    note = cell(ws2, r, 10)
    key = str(subsub) if subsub else (str(sub) if sub else (str(main) if main else None))
    if key is None or q_label not in QMAP:
        continue
    qfield = QMAP[q_label]
    if qfield in ('q3', 'q4'):
        continue  # V1 rule: do not seed Q3/Q4
    quarterly.setdefault(key, {})
    quarterly[key][qfield] = {'actual': actual, 'score': score, 'note': note}

# ── 3. คาดการณ์: forecast final result + score ──
ws3 = wb['3.คาดการณ์']
forecast = {}
for r in range(4, ws3.max_row + 1):
    kpi_id = cell(ws3, r, 2)
    fc_result = cell(ws3, r, 5)
    fc_score = cell(ws3, r, 6)
    fc_note = cell(ws3, r, 8)
    if kpi_id is None:
        continue
    forecast[str(kpi_id)] = {'result': fc_result, 'score': fc_score, 'note': fc_note}

# ── 4. Action_Plan: activity/milestone steps (1.3, 2.7, 2.8) ──
ws4 = wb['4.Action_Plan']
activities = {}
for r in range(3, ws4.max_row + 1):
    main = cell(ws4, r, 2)
    sub = cell(ws4, r, 3)
    subsub = cell(ws4, r, 4)
    note = cell(ws4, r, 14)
    key = str(subsub) if subsub else (str(sub) if sub else (str(main) if main else None))
    if key is None:
        continue
    steps = []
    for c in range(6, 14):
        v = cell(ws4, r, c)
        if v:
            steps.append(v)
    activities[key] = {'steps': steps, 'note': note}

# ── 5. หมายเหตุหักคะแนน: deduction conditions ──
ws5 = wb['5. หมายเหตุหักคะแนน']
deductions = []
for r in range(4, ws5.max_row + 1):
    kpi_id = cell(ws5, r, 2)
    reason = cell(ws5, r, 3)
    pts = cell(ws5, r, 4)
    status = cell(ws5, r, 5)
    if kpi_id is None:
        continue
    deductions.append({'kpi': str(kpi_id), 'reason': reason, 'points': pts, 'status': status})

# ── 6. Meeting_Assignment: real board follow-up items ──
ws6 = wb['6. Meeting_Assignment']
assignments = []
for r in range(4, ws6.max_row + 1):
    main = cell(ws6, r, 2)
    sub = cell(ws6, r, 3)
    subsub = cell(ws6, r, 4)
    note = cell(ws6, r, 5)
    key = str(subsub) if subsub else (str(sub) if sub else (str(main) if main else None))
    if key is None or note is None:
        continue
    assignments.append({'kpi': key, 'note': note})

# ── 7. Executive_Owner: owner → KPI responsibility map ──
ws7 = wb['7. Executive_Owner']
exec_owners = []
for r in range(4, ws7.max_row + 1):
    group = cell(ws7, r, 2)
    person = cell(ws7, r, 5)
    role = cell(ws7, r, 17)
    if person is None:
        continue
    kpi_list = []
    for c in range(6, 17):
        v = cell(ws7, r, c)
        if v:
            kpi_list.append(str(v))
    exec_owners.append({'group': group, 'person': person, 'role': role, 'kpis': kpi_list})

# ── 9. การเชื่อมโยง: Home cross-filter design map ──
ws9 = wb['9. การเชื่อมโยง']
linkage = []
for r in range(4, ws9.max_row + 1):
    box = cell(ws9, r, 2)
    goto = cell(ws9, r, 3)
    filt = cell(ws9, r, 4)
    present = cell(ws9, r, 5)
    origin = cell(ws9, r, 6)
    if box is None:
        continue
    linkage.append({'box': box, 'goto': goto, 'filter': filt, 'presents': present, 'origin': origin})

# ── 9. ผลรายเดือน: raw monthly actuals — V1 rule: only ต.ค.-มี.ค. (Q1+Q2), never เม.ย.-ก.ย. ──
# Quarter Result / Cumulative Result are NOT pre-computed here — the app derives them
# from these raw months at render time (Detail page §5A), never a new scoring formula.
ws9b = wb['9. ผลรายเดือน']
MONTH_COLS = [('m1', 7), ('m2', 8), ('m3', 9), ('m4', 10), ('m5', 11), ('m6', 12)]  # G-L = ต.ค.-มี.ค.
monthly = {}
for r in range(3, ws9b.max_row + 1):
    main = cell(ws9b, r, 2)
    sub = cell(ws9b, r, 3)
    subsub = cell(ws9b, r, 4)
    unit = cell(ws9b, r, 5)
    label = cell(ws9b, r, 6)
    key = str(subsub) if subsub else (str(sub) if sub else (str(main) if main else None))
    if key is None:
        continue
    months = {}
    for mkey, c in MONTH_COLS:
        v = cell(ws9b, r, c)
        if v is not None:
            months[mkey] = v
    if not months:
        continue
    monthly.setdefault(key, {'unit': unit, 'components': []})
    monthly[key]['components'].append({'label': label, 'months': months})

out = {
    'meta': {
        'source': 'MOU69_Claude.xlsx (restructured 2026-08-27)',
        'generated': datetime.now().isoformat(),
        'rule': 'V1 seeds Q1+Q2 only. Q3/Q4 intentionally dropped even though present in source.',
    },
    'kpis': kpis,
    'quarterly': quarterly,
    'forecast': forecast,
    'activities': activities,
    'deductions': deductions,
    'assignments': assignments,
    'execOwners': exec_owners,
    'linkage': linkage,
    'monthly': monthly,
}

with open('mou_data.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

with open('mou_data.js', 'w', encoding='utf-8') as f:
    f.write('// AUTO-GENERATED by extract_data.py — do not hand-edit. Re-run the script instead.\n')
    f.write('const MOU_DATA = ')
    f.write(json.dumps(out, ensure_ascii=False, indent=2))
    f.write(';\n')

print(f"KPIs: {len(kpis)}")
print(f"Total weight: {sum(k['weight'] for k in kpis.values() if k['weight'])}")
print("Wrote mou_data.json + mou_data.js")
